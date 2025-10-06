import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from k_anonymity_implementation import run_k_anonymity_analysis

def create_excel_report(filename='k_anonymity_report.xlsx'):
    """
    Create comprehensive Excel report with k-anonymity analysis results.
    Includes multiple sheets with data, metrics, and embedded charts.
    """
    print("Generating K-Anonymity Excel Report...")
    print("=" * 80)

    # Get results from k-anonymity implementation
    print("\n[1/7] Running k-anonymity analysis...")
    results = run_k_anonymity_analysis(print_results=False, return_results=True)

    # Create Excel writer
    print("[2/7] Creating Excel workbook...")
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Sheet 1: Original Dataset
    print("[3/7] Writing original dataset...")
    df_original = results['original_df']
    df_original.to_excel(writer, sheet_name='Original Dataset', index=False)

    # Sheet 2: K=3 Anonymized
    print("[4/7] Writing k=3 anonymized dataset...")
    df_k3 = results['anonymized_dfs'][3]
    df_k3.to_excel(writer, sheet_name='K=3 Anonymized', index=False)

    # Sheet 3: K=5 Anonymized
    print("[5/7] Writing k=5 anonymized dataset...")
    df_k5 = results['anonymized_dfs'][5]
    df_k5.to_excel(writer, sheet_name='K=5 Anonymized', index=False)

    # Sheet 4: Metrics Comparison
    print("[6/7] Creating metrics comparison sheet...")
    metrics_data = []

    for k in results['k_values']:
        m = results['metrics'][k]
        metrics_data.append({
            'K Value': k,
            'Re-identification Probability (%)': m['reident_prob'] * 100,
            'Privacy Gain (%)': m['privacy_gain'] * 100,
            'Min Equivalence Class Size': m['min_size'],
            'Discernibility Cost (Original)': m['discern_orig'],
            'Discernibility Cost (Anonymized)': m['discern_anon'],
            'Information Loss Increase (%)': ((m['discern_anon'] - m['discern_orig']) / m['discern_orig'] * 100),
            'Age Precision Loss (%)': m['precision_loss'].get('Age', 0) * 100,
            'ZipCode Precision Loss (%)': m['precision_loss'].get('ZipCode', 0) * 100,
            'K-Anonymous': 'YES' if m['is_anon'] else 'NO'
        })

    df_metrics = pd.DataFrame(metrics_data)
    df_metrics.to_excel(writer, sheet_name='Metrics Comparison', index=False)

    # Sheet 5: Equivalence Classes
    print("[7/7] Writing equivalence classes analysis...")
    eq_data = []

    for k in results['k_values']:
        eq_classes = results['equivalence_classes'][k]
        for i, (qi_combo, indices) in enumerate(eq_classes.items(), 1):
            eq_data.append({
                'K Value': k,
                'Class Number': i,
                'ZipCode': qi_combo[0],
                'Age': qi_combo[1],
                'Gender': qi_combo[2],
                'Record Count': len(indices)
            })

    df_eq = pd.DataFrame(eq_data)
    df_eq.to_excel(writer, sheet_name='Equivalence Classes', index=False)

    # Save the workbook
    writer.close()

    # Now add charts and formatting
    print("\nAdding charts and formatting...")
    wb = openpyxl.load_workbook(filename)

    # Format Original Dataset sheet
    format_data_sheet(wb['Original Dataset'], "Original Medical Dataset")

    # Format K=3 Anonymized sheet
    format_data_sheet(wb['K=3 Anonymized'], "K=3 Anonymized Dataset")

    # Format K=5 Anonymized sheet
    format_data_sheet(wb['K=5 Anonymized'], "K=5 Anonymized Dataset")

    # Format and add charts to Metrics sheet
    format_metrics_sheet(wb['Metrics Comparison'])

    # Format Equivalence Classes sheet
    format_eq_classes_sheet(wb['Equivalence Classes'])

    # Save formatted workbook
    wb.save(filename)

    print("\n" + "=" * 80)
    print(f"✓ Excel report generated successfully: {filename}")
    print("=" * 80)

    print("\nWorkbook contains:")
    print("  • Sheet 1: Original Dataset (15 records)")
    print("  • Sheet 2: K=3 Anonymized Dataset")
    print("  • Sheet 3: K=5 Anonymized Dataset")
    print("  • Sheet 4: Metrics Comparison (with embedded charts)")
    print("  • Sheet 5: Equivalence Classes Analysis")
    print("\nCharts included:")
    print("  • Privacy Protection Levels")
    print("  • Re-identification Probability")
    print("  • Information Loss Comparison")
    print("  • Precision Loss by Attribute")

def format_data_sheet(ws, title):
    """Apply formatting to data sheets."""
    # Set title
    ws.insert_rows(1)
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Format header row
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def format_metrics_sheet(ws):
    """Format metrics sheet and add charts."""
    # Format header
    ws.insert_rows(1)
    ws['A1'] = 'K-Anonymity Metrics Comparison'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Format data header
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Format data cells
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 32
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 28
    ws.column_dimensions['J'].width = 15

    # Add Chart 1: Privacy Gain
    chart1 = BarChart()
    chart1.type = 'col'
    chart1.title = 'Privacy Protection Level'
    chart1.y_axis.title = 'Privacy Gain (%)'
    chart1.x_axis.title = 'K Value'

    data = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 10
    chart1.width = 15

    ws.add_chart(chart1, 'L3')

    # Add Chart 2: Re-identification Probability
    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = 'Re-identification Probability'
    chart2.y_axis.title = 'Probability (%)'
    chart2.x_axis.title = 'K Value'

    data2 = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 10
    chart2.width = 15

    ws.add_chart(chart2, 'L20')

    # Add Chart 3: Information Loss
    chart3 = LineChart()
    chart3.title = 'Information Loss Comparison'
    chart3.y_axis.title = 'Discernibility Cost'
    chart3.x_axis.title = 'K Value'

    data3_orig = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
    data3_anon = Reference(ws, min_col=6, min_row=2, max_row=ws.max_row)
    chart3.add_data(data3_orig, titles_from_data=True)
    chart3.add_data(data3_anon, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.height = 10
    chart3.width = 15

    ws.add_chart(chart3, 'V3')

    # Add Chart 4: Precision Loss
    chart4 = BarChart()
    chart4.type = 'col'
    chart4.grouping = 'clustered'
    chart4.title = 'Attribute Precision Loss'
    chart4.y_axis.title = 'Precision Loss (%)'
    chart4.x_axis.title = 'K Value'

    data4_age = Reference(ws, min_col=8, min_row=2, max_row=ws.max_row)
    data4_zip = Reference(ws, min_col=9, min_row=2, max_row=ws.max_row)
    chart4.add_data(data4_age, titles_from_data=True)
    chart4.add_data(data4_zip, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.height = 10
    chart4.width = 15

    ws.add_chart(chart4, 'V20')

def format_eq_classes_sheet(ws):
    """Format equivalence classes sheet."""
    # Format header
    ws.insert_rows(1)
    ws['A1'] = 'Equivalence Classes Analysis'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Format data header
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    # Add borders and center align
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

if __name__ == "__main__":
    create_excel_report()
    print("\nYou can now open 'k_anonymity_report.xlsx' in Excel or Google Sheets!")
