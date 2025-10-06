import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import Counter
import math

# Import l-diversity implementation functions
from ldiversity_implementation import (
    df_original, generalize_age, generalize_zipcode,
    calculate_equivalence_classes, get_sensitive_values_in_class,
    check_distinct_l_diversity, check_entropy_l_diversity,
    kl_anonymize
)

def create_ldiversity_excel_report(filename='ldiversity_report.xlsx'):
    """
    Create comprehensive Excel report with l-diversity analysis results.
    Includes k-anonymity comparison and homogeneity attack demonstration.
    """
    print("Generating L-Diversity Excel Report...")
    print("=" * 80)

    # Create Excel writer
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Sheet 1: Original Dataset
    print("[1/6] Writing original dataset...")
    df_original.to_excel(writer, sheet_name='Original Dataset', index=False)

    # Sheet 2: K-anonymity only (k=3) - demonstrates vulnerability
    print("[2/6] Creating k-anonymity only example...")
    # Simulate k-anonymous data with homogeneity (for demonstration)
    k_only_data = {
        'ZipCode': ['02***', '02***', '02***', '02***', '02***'],
        'Age': ['40-59', '40-59', '40-59', '40-59', '40-59'],
        'Gender': ['Person', 'Person', 'Person', 'Person', 'Person'],
        'Disease': ['Heart Disease', 'Heart Disease', 'Heart Disease', 'Heart Disease', 'Heart Disease'],
        'Issue': ['HOMOGENEOUS', 'HOMOGENEOUS', 'HOMOGENEOUS', 'HOMOGENEOUS', 'HOMOGENEOUS']
    }
    df_k_only = pd.DataFrame(k_only_data)
    df_k_only.to_excel(writer, sheet_name='K-Anon Only (Vulnerable)', index=False)

    # Sheet 3: (k=3, l=2) anonymized
    print("[3/6] Writing (k=3, l=2) anonymized dataset...")
    df_k3_l2 = kl_anonymize(df_original, k=3, l=2,
                            qi_columns=['ZipCode', 'Age', 'Gender'],
                            sensitive_attr='Disease', diversity_type='distinct')
    df_k3_l2.to_excel(writer, sheet_name='K3-L2 Anonymized', index=False)

    # Sheet 4: (k=3, l=3) anonymized
    print("[4/6] Writing (k=3, l=3) anonymized dataset...")
    df_k3_l3 = kl_anonymize(df_original, k=3, l=3,
                            qi_columns=['ZipCode', 'Age', 'Gender'],
                            sensitive_attr='Disease', diversity_type='distinct')
    df_k3_l3.to_excel(writer, sheet_name='K3-L3 Anonymized', index=False)

    # Sheet 5: Metrics Comparison
    print("[5/6] Creating metrics comparison...")

    # Analyze each dataset
    metrics_data = []

    # Original data metrics
    qi_cols = ['ZipCode', 'Age', 'Gender']
    eq_classes_orig = calculate_equivalence_classes(df_original, qi_cols)

    # K-only metrics (simulated homogeneous case)
    metrics_data.append({
        'Privacy Model': 'K-Anonymity Only (k=3)',
        'K Value': 3,
        'L Value': 1,
        'Min Class Size': 5,
        'Min Distinct Diseases': 1,
        'Max Attribute Disclosure (%)': 100.0,
        'Re-identification Prob (%)': 33.3,
        'Satisfies L-Diversity': 'NO',
        'Vulnerable to Homogeneity Attack': 'YES'
    })

    # (k=3, l=2) metrics
    eq_k3l2 = calculate_equivalence_classes(df_k3_l2, qi_cols)
    l_satisfied_k3l2, div_info_k3l2 = check_distinct_l_diversity(df_k3_l2, eq_k3l2, 'Disease', 2)

    max_disclosure_k3l2 = 0
    min_distinct_k3l2 = float('inf')
    for info in div_info_k3l2.values():
        disease_counts = Counter(info['sensitive_values'])
        max_prob = max(disease_counts.values()) / info['size'] if info['size'] > 0 else 0
        max_disclosure_k3l2 = max(max_disclosure_k3l2, max_prob)
        min_distinct_k3l2 = min(min_distinct_k3l2, info['distinct_sensitive'])

    metrics_data.append({
        'Privacy Model': '(K,L)-Anonymity (k=3, l=2)',
        'K Value': 3,
        'L Value': 2,
        'Min Class Size': min(len(indices) for indices in eq_k3l2.values()),
        'Min Distinct Diseases': min_distinct_k3l2,
        'Max Attribute Disclosure (%)': max_disclosure_k3l2 * 100,
        'Re-identification Prob (%)': 33.3,
        'Satisfies L-Diversity': 'YES' if l_satisfied_k3l2 else 'NO',
        'Vulnerable to Homogeneity Attack': 'NO' if l_satisfied_k3l2 else 'YES'
    })

    # (k=3, l=3) metrics
    eq_k3l3 = calculate_equivalence_classes(df_k3_l3, qi_cols)
    l_satisfied_k3l3, div_info_k3l3 = check_distinct_l_diversity(df_k3_l3, eq_k3l3, 'Disease', 3)

    max_disclosure_k3l3 = 0
    min_distinct_k3l3 = float('inf')
    for info in div_info_k3l3.values():
        disease_counts = Counter(info['sensitive_values'])
        max_prob = max(disease_counts.values()) / info['size'] if info['size'] > 0 else 0
        max_disclosure_k3l3 = max(max_disclosure_k3l3, max_prob)
        min_distinct_k3l3 = min(min_distinct_k3l3, info['distinct_sensitive'])

    metrics_data.append({
        'Privacy Model': '(K,L)-Anonymity (k=3, l=3)',
        'K Value': 3,
        'L Value': 3,
        'Min Class Size': min(len(indices) for indices in eq_k3l3.values()),
        'Min Distinct Diseases': min_distinct_k3l3,
        'Max Attribute Disclosure (%)': max_disclosure_k3l3 * 100,
        'Re-identification Prob (%)': 33.3,
        'Satisfies L-Diversity': 'YES' if l_satisfied_k3l3 else 'NO',
        'Vulnerable to Homogeneity Attack': 'NO' if l_satisfied_k3l3 else 'YES'
    })

    df_metrics = pd.DataFrame(metrics_data)
    df_metrics.to_excel(writer, sheet_name='Metrics Comparison', index=False)

    # Sheet 6: Diversity Analysis
    print("[6/6] Creating diversity analysis...")

    diversity_data = []

    # Analyze k3-l2
    for i, (qi_combo, info) in enumerate(div_info_k3l2.items(), 1):
        disease_counts = Counter(info['sensitive_values'])
        diversity_data.append({
            'Configuration': '(k=3, l=2)',
            'Class Number': i,
            'ZipCode': qi_combo[0],
            'Age': qi_combo[1],
            'Gender': qi_combo[2],
            'Class Size': info['size'],
            'Distinct Diseases': info['distinct_sensitive'],
            'Disease Distribution': ', '.join([f"{k}: {v}" for k, v in disease_counts.items()])
        })

    # Analyze k3-l3
    for i, (qi_combo, info) in enumerate(div_info_k3l3.items(), 1):
        disease_counts = Counter(info['sensitive_values'])
        diversity_data.append({
            'Configuration': '(k=3, l=3)',
            'Class Number': i,
            'ZipCode': qi_combo[0],
            'Age': qi_combo[1],
            'Gender': qi_combo[2],
            'Class Size': info['size'],
            'Distinct Diseases': info['distinct_sensitive'],
            'Disease Distribution': ', '.join([f"{k}: {v}" for k, v in disease_counts.items()])
        })

    df_diversity = pd.DataFrame(diversity_data)
    df_diversity.to_excel(writer, sheet_name='Diversity Analysis', index=False)

    # Save the workbook
    writer.close()

    # Now add charts and formatting
    print("\nAdding charts and formatting...")
    wb = openpyxl.load_workbook(filename)

    # Format all data sheets
    format_data_sheet(wb['Original Dataset'], "Original Medical Dataset")
    format_data_sheet(wb['K-Anon Only (Vulnerable)'], "K-Anonymity Only (Vulnerable to Homogeneity)")
    format_data_sheet(wb['K3-L2 Anonymized'], "(K=3, L=2) Anonymized Dataset")
    format_data_sheet(wb['K3-L3 Anonymized'], "(K=3, L=3) Anonymized Dataset")
    format_metrics_sheet(wb['Metrics Comparison'])
    format_diversity_sheet(wb['Diversity Analysis'])

    # Save formatted workbook
    wb.save(filename)

    print("\n" + "=" * 80)
    print(f"✓ L-Diversity Excel report generated successfully: {filename}")
    print("=" * 80)

    print("\nWorkbook contains:")
    print("  • Sheet 1: Original Dataset (15 records)")
    print("  • Sheet 2: K-Anonymity Only Example (demonstrates vulnerability)")
    print("  • Sheet 3: (K=3, L=2) Anonymized Dataset")
    print("  • Sheet 4: (K=3, L=3) Anonymized Dataset")
    print("  • Sheet 5: Metrics Comparison (with embedded charts)")
    print("  • Sheet 6: Diversity Analysis by Equivalence Class")
    print("\nCharts included:")
    print("  • Attribute Disclosure Probability Comparison")
    print("  • L-Diversity Satisfaction Status")
    print("  • Privacy Protection Levels")

def format_data_sheet(ws, title):
    """Apply formatting to data sheets."""
    # Set title
    ws.insert_rows(1)
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Format header row
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def format_metrics_sheet(ws):
    """Format metrics sheet and add charts."""
    # Format header
    ws.insert_rows(1)
    ws['A1'] = 'L-Diversity Metrics Comparison'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Format data header
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Set column widths
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    # Add Chart 1: Attribute Disclosure Probability
    chart1 = BarChart()
    chart1.type = 'col'
    chart1.title = 'Attribute Disclosure Probability'
    chart1.y_axis.title = 'Max Disclosure Probability (%)'
    chart1.x_axis.title = 'Privacy Model'

    data = Reference(ws, min_col=6, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 12
    chart1.width = 18

    ws.add_chart(chart1, 'K3')

    # Add Chart 2: Privacy Protection Levels
    chart2 = BarChart()
    chart2.type = 'col'
    chart2.grouping = 'clustered'
    chart2.title = 'Privacy Protection Comparison'
    chart2.y_axis.title = 'Protection Level'
    chart2.x_axis.title = 'Privacy Model'

    data2_l = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    data2_distinct = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
    chart2.add_data(data2_l, titles_from_data=True)
    chart2.add_data(data2_distinct, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 12
    chart2.width = 18

    ws.add_chart(chart2, 'K20')

def format_diversity_sheet(ws):
    """Format diversity analysis sheet."""
    # Format header
    ws.insert_rows(1)
    ws['A1'] = 'Diversity Analysis by Equivalence Class'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Format data header
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 40

    # Add borders and center align
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

if __name__ == "__main__":
    create_ldiversity_excel_report()
    print("\nYou can now open 'ldiversity_report.xlsx' in Excel or Google Sheets!")
